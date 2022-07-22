# EMULATE
# Author: Kevin Hines
#Date: 07/12/2022
import serial 
import time
from simple_pid import PID
from openpyxl import Workbook

#Open and excel workbook and create a new sheet
xlBook = Workbook()
sheetName = 'Cal Sheet'
xlSheet = xlBook.create_sheet(sheetName, 0)

#Setup column headers for the data
portHeader = xlSheet.cell(row=1, column=1, value="Port")
countHeader = xlSheet.cell(row=1, column=2, value="Count")
PressureHeader = xlSheet.cell(row=1, column=3, value="Pressure")
rowCount = 2

#Ensures the user has to manual enter dual regulator control and prompts for confirmation
print("Please enter CoolTerm and enable Dual Regulator Control before continuing. ")
cont = str(input("Once Dual Regulator Control is enabled and you have disconnected CoolTerm, type 'y' to continue: "))

if cont != 'y':
    exit()

#Gathers COM ports for devices
print("Please format the address as COM followed by the specific number for that device.")

gaugeAdd = input("Port number for Pressure Gauge: ")
zoeAdd = input("Port number for Zoe: ")

gaugePort = serial.Serial(
    port = gaugeAdd,
    baudrate = 9600,
    bytesize=serial.EIGHTBITS,
    parity=serial.PARITY_NONE,
    stopbits=serial.STOPBITS_ONE,
    timeout=1   
)

zoePort = serial.Serial(
    port = zoeAdd,
    baudrate = 57600,
    bytesize=serial.EIGHTBITS,
    parity=serial.PARITY_NONE,
    stopbits=serial.STOPBITS_ONE,
    timeout=1   
)

###
# Uses the user input to send a message to the Zoe about which port is being calibrated

#Input:
# - desiredPort: the port selected by the user.
###
def setPort(desiredPort):
    zoePort.write(b'V')
    if desiredPort == 1:
        zoePort.write(b'1')
    if desiredPort == 2:
        zoePort.write(b'2')
    if desiredPort == 3:
        zoePort.write(b'3')
    if desiredPort == 4:
        zoePort.write(b'4')
    if desiredPort == 5:
        zoePort.write(b'5')
        
###
# This function prompts the user for the calibration set values

#Input:
# -rowNum: the number of the next empty row in the excel sheet which the cal values are stored in
# -port: the port number being calibrated
###
def portSetPoints():
    print("Input all the calibration values for this port. If you mistype a number, type the letter a and press enter to retry.")
    while True:
        setPointOne = input("Input the first pressure set point: ")
        try:
            setPointOne = float(setPointOne)
            break
        except:
            print("Invalid input, try again")

    while True:
        setPointTwo = input("Input the second pressure set point: ")
        try:
            setPointTwo = float(setPointTwo)
            break
        except:
            print("Invalid input, try again")
    #Currently commented out as we need to figure out how to have a 0 set point that works
    #while True:
        #setPointThree = input("Input the third pressure set point: ")
        #try:
            #setPointThree = float(setPointThree)
            #break
        #except:
            #print("Invalid input, try again")

    points = [setPointOne, setPointTwo]
    return points

###
# runCal uses parameters from the calPort function to run a PID loop for reaching a set goal
# pressure value from the baseline. It does this PID loop in two steps with the first one being faster
# and less precise. After going through the first loop and reaching within a set tolerance,
# it will move onto the more precise loop to get within specification. 

#Input:
# - goal: the desired pressure value
# - min: the minimum pressure value that the PID can reach.
# - max: the maximum pressure value that the PID can reach.
# - rowNum: the current row index on the excel sheet used to record final count values.
###
def runCal(goal, min, max, rowNum):
    waitTime = 20
    runCount = 0
    tolerance = 0.1

    print(goal)
    pid = PID(0.72, 0.07, 0.05, setpoint=goal)
    pid.output_limits = (min, max) 

    print("Returning to baseline...")
    zoePort.write(b'2300\n')
    time.sleep(3)

    gaugePort.write(b"255:R:MRMD:1\r\n")
    data = str(gaugePort.readline())

    measuredPressure = float(data[13:21])

    while runCount < 2:
        while measuredPressure > (goal + tolerance) or measuredPressure < (goal-tolerance):
            gaugePort.write(b"255:R:MRMD:1\r\n")
            data = str(gaugePort.readline())
            measuredPressure = float(data[13:21])
            print("Measured Pressure:", measuredPressure)

            output = (int(pid(measuredPressure)*10))
            message = bytes((str(output)+'\n'), 'utf-8')
            print("New Value:", output)

            print("Returning to baseline...")
            zoePort.write(b'2300\n')
            time.sleep(3)
            
            print("Trying new value...")
            zoePort.write(message)
            time.sleep(waitTime)

        runCount+=1
        waitTime = 70
        tolerance = 0.02
        pid.Kp = 0.68
        pid.Ki = 0.03
        pid.Kd = 0.08

    portLocation = 'A' + str(rowNum)
    countLocation = 'B' + str(rowNum)
    pressureLocation = 'C' + str(rowNum)
    
    try:
        xlSheet[portLocation] = portNum
        xlSheet[countLocation] = message
        xlSheet[pressureLocation] = measuredPressure
        rowNum += 1
    except:
        print("No values saved")
    return rowNum

###
# Function receives a port and gathers the calibration set points. It then goes through
# each set point and calls the runCal function to auto calibrate for each pressure set point.
# It also differentiates between vacuum and the rest of the ports to approriately set min
# and max for the pid

#Input:
# -rowNum: the number of the next empty row in the excel sheet which the cal values are stored in
# -port: the port number being calibrated
###
def calPort(rowNum, port):
    setPoints = portSetPoints()
    for i in setPoints:
        if port == 5:
            rowNum = runCal(i, 60, 207, rowNum)
        else:
            rowNum = runCal(i, 204, 320, rowNum)
    return rowNum

while True:
    #While running, ask the user what port to calibrate and then call calibration functions for that port
    portNum = int(input("Input the port to calibrate (1 = TI, 2 = TO, 3 = BI, 4 = BO, 5 = V, 6 = Quit): "))
    if portNum == 6:
        break

    setPort(portNum)
    rowCount = calPort(rowCount, portNum)
    #Writes control + Q to Zoe to cancel out of that port once calibration is done
    zoePort.write(bytes(chr(17), 'utf-8'))

print("Communcation Closed")

gaugePort.close()
zoePort.close()

#saves excel sheet with count values
saveName = sheetName + '.csv'
xlBook.save(saveName)
